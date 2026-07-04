from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, F, Count
from accounts.decorators import petugas_required
from produk.models import Produk
from transaksi.models import Transaksi, DetailTransaksi


def _get_filtered_qs(request):
    """Ambil rentang tanggal dari query param dan siapkan queryset transaksi yang sudah difilter."""
    tanggal_awal = request.GET.get('tanggal_awal', '').strip()
    tanggal_akhir = request.GET.get('tanggal_akhir', '').strip()

    transaksi_qs = Transaksi.objects.all()
    if tanggal_awal:
        transaksi_qs = transaksi_qs.filter(tanggal__date__gte=tanggal_awal)
    if tanggal_akhir:
        transaksi_qs = transaksi_qs.filter(tanggal__date__lte=tanggal_akhir)

    return transaksi_qs, tanggal_awal, tanggal_akhir


@petugas_required
def laporan_index(request):

    transaksi_qs, tanggal_awal, tanggal_akhir = _get_filtered_qs(request)

    # =====================
    # A. RINGKASAN INVENTORY (posisi stok saat ini, tidak terpengaruh filter tanggal)
    # =====================
    total_item = Produk.objects.aggregate(
        total=Sum('stok')
    )['total'] or 0

    total_nilai_inventory = Produk.objects.aggregate(
        total=Sum(F('stok') * F('harga_beli'))
    )['total'] or 0


    # =====================
    # B. BARANG MASUK (mengikuti filter tanggal)
    # =====================
    transaksi_masuk = transaksi_qs.filter(jenis_transaksi='MASUK')

    total_transaksi_masuk = transaksi_masuk.count()

    total_unit_masuk = DetailTransaksi.objects.filter(
        transaksi__in=transaksi_masuk
    ).aggregate(total=Sum('jumlah'))['total'] or 0

    total_nilai_masuk = DetailTransaksi.objects.filter(
        transaksi__in=transaksi_masuk
    ).aggregate(
        total=Sum(F('jumlah') * F('produk__harga_beli'))
    )['total'] or 0


    # =====================
    # C. BARANG KELUAR (mengikuti filter tanggal)
    # =====================
    transaksi_keluar = transaksi_qs.filter(jenis_transaksi='KELUAR')

    total_transaksi_keluar = transaksi_keluar.count()

    total_unit_keluar = DetailTransaksi.objects.filter(
        transaksi__in=transaksi_keluar
    ).aggregate(total=Sum('jumlah'))['total'] or 0

    total_nilai_keluar = DetailTransaksi.objects.filter(
        transaksi__in=transaksi_keluar
    ).aggregate(
        total=Sum(F('jumlah') * F('produk__harga_beli'))
    )['total'] or 0


    # =====================
    # D. NILAI INVENTORY PER KATEGORI (posisi stok saat ini)
    # =====================
    nilai_per_kategori = (
        Produk.objects
        .filter(is_active=True)
        .values('kategori__nama_kategori')
        .annotate(
            total_unit=Sum('stok'),
            total_nilai=Sum(F('stok') * F('harga_beli'))
        )
        .order_by('-total_nilai')
    )


    # =====================
    # E. RINGKASAN PRODUK (mengikuti filter tanggal, dibatasi 20 teratas)
    # =====================
    ringkasan_produk = (
        DetailTransaksi.objects
        .filter(transaksi__in=transaksi_qs)
        .values('produk__nama_produk')
        .annotate(
            jumlah_transaksi=Count('id'),
            total_unit=Sum('jumlah'),
            total_nilai=Sum(F('jumlah') * F('produk__harga_beli'))
        )
        .order_by('-total_nilai')[:20]
    )


    # =====================
    # F. PRODUK DENGAN NILAI TERTINGGI (TOP 10, posisi stok saat ini)
    # =====================
    produk_tertinggi = (
        Produk.objects
        .filter(is_active=True)
        .annotate(
            total_nilai=F('stok') * F('harga_beli')
        )
        .order_by('-total_nilai')[:10]
    )


    context = {
        'total_item': total_item,
        'total_nilai_inventory': total_nilai_inventory,

        'total_transaksi_masuk': total_transaksi_masuk,
        'total_unit_masuk': total_unit_masuk,
        'total_nilai_masuk': total_nilai_masuk,

        'total_transaksi_keluar': total_transaksi_keluar,
        'total_unit_keluar': total_unit_keluar,
        'total_nilai_keluar': total_nilai_keluar,

        'nilai_per_kategori': nilai_per_kategori,
        'ringkasan_produk': ringkasan_produk,
        'produk_tertinggi': produk_tertinggi,

        'tanggal_awal': tanggal_awal,
        'tanggal_akhir': tanggal_akhir,
    }

    return render(request, 'laporan/index.html', context)


@petugas_required
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.utils.timezone import now

    transaksi_qs, tanggal_awal, tanggal_akhir = _get_filtered_qs(request)

    wb = Workbook()

    # ===== SHEET 1: RINGKASAN INVENTORY =====
    ws1 = wb.active
    ws1.title = "Ringkasan Inventory"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    ws1.append(["Laporan Inventory - Toko Borneo Ponsel"])
    ws1.append([f"Dicetak pada: {now().strftime('%d-%m-%Y %H:%M')}"])
    if tanggal_awal or tanggal_akhir:
        ws1.append([f"Periode transaksi: {tanggal_awal or 'awal'} s/d {tanggal_akhir or 'sekarang'}"])
    ws1.append([])

    ws1.append(["Kode Produk", "Nama Produk", "Kategori", "Stok", "Stok Minimum",
                "Harga Beli", "Harga Jual", "Nilai Inventory"])
    header_row = ws1.max_row
    for cell in ws1[header_row]:
        cell.font = header_font
        cell.fill = header_fill

    produk_qs = Produk.objects.filter(is_active=True).select_related('kategori')
    for p in produk_qs:
        ws1.append([
            p.kode_produk, p.nama_produk, p.kategori.nama_kategori,
            p.stok, p.stok_minimum, float(p.harga_beli), float(p.harga_jual),
            float(p.stok * p.harga_beli),
        ])

    for col in ws1.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws1.column_dimensions[col[0].column_letter].width = max_len + 3

    # ===== SHEET 2: RIWAYAT TRANSAKSI (mengikuti filter tanggal) =====
    ws2 = wb.create_sheet("Riwayat Transaksi")
    if tanggal_awal or tanggal_akhir:
        ws2.append([f"Periode: {tanggal_awal or 'awal'} s/d {tanggal_akhir or 'sekarang'}"])
        ws2.append([])
    ws2.append(["Tanggal", "Jenis", "Produk", "Jumlah", "Supplier", "Keterangan", "User"])
    header_row2 = ws2.max_row
    for cell in ws2[header_row2]:
        cell.font = header_font
        cell.fill = header_fill

    detail_qs = DetailTransaksi.objects.filter(
        transaksi__in=transaksi_qs
    ).select_related(
        'transaksi', 'produk', 'transaksi__user', 'transaksi__supplier'
    ).order_by('-transaksi__tanggal')

    for d in detail_qs:
        ws2.append([
            d.transaksi.tanggal.strftime('%d-%m-%Y %H:%M'),
            d.transaksi.get_jenis_transaksi_display(),
            d.produk.nama_produk,
            d.jumlah,
            d.transaksi.supplier.nama_supplier if d.transaksi.supplier else '-',
            d.transaksi.keterangan or '-',
            str(d.transaksi.user) if d.transaksi.user else '-',
        ])

    for col in ws2.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws2.column_dimensions[col[0].column_letter].width = max_len + 3

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"laporan-inventory-{now().strftime('%Y%m%d-%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
